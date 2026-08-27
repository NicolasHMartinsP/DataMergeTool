import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from collections import defaultdict
import config
from models.entity import Entity
import gspread
from google.oauth2.service_account import Credentials

class ExcelService:
    def __init__(self):
        self.base_sheet_mode_1 = getattr(config, 'MASTER_MODE_1_SHEET', "Mode_1_Sheet")
        self.base_sheet_mode_2 = getattr(config, 'MASTER_MODE_2_SHEET', "Mode_2_Sheet")
        
        self.allowed_mode_1_sheets = getattr(config, 'ALLOWED_MODE_1_SHEETS', [])
        self.allowed_mode_2_sheets = getattr(config, 'ALLOWED_MODE_2_SHEETS', [])
        
        self.files_to_scan = []
        self.mapping = []
        self.workbooks = {}

    def _connect_cloud(self):
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly"
        ]
        
        credentials_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "credentials.json")
        
        if not os.path.exists(credentials_path):
            raise Exception("credentials.json not found in the project root! Ensure you have Google Service Account credentials.")
            
        credentials = Credentials.from_service_account_file(credentials_path, scopes=scopes)
        client = gspread.authorize(credentials)
        return client

    def read_master_records(self, mode: int):
        try:
            print("\n>>> CONNECTING TO CLOUD SERVICES...")
            client = self._connect_cloud()
            
            cloud_spreadsheet = client.open_by_url(config.MASTER_SPREADSHEET_URL)
            sheet_name = self.base_sheet_mode_1 if mode == 1 else self.base_sheet_mode_2
            
            try:
                sheet = cloud_spreadsheet.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                sheet = cloud_spreadsheet.sheet1
                
            print(f">>> DOWNLOAD OF MASTER DATABASE ('{sheet_name}') COMPLETED!\n")
            
            data = sheet.get_all_values()
            
            if not data:
                raise Exception("The online spreadsheet is empty.")
                
            header = [str(h).strip().upper() for h in data[0]]
            
            col_id = -1
            col_name = -1
            
            for i, h in enumerate(header):
                words = h.replace("_", " ").split()
                if "ID" in words or h == "ID":
                    col_id = i
                if h == "NOME":
                    col_name = i
                    
            if col_id == -1 or col_name == -1:
                raise Exception("Could not find 'ID' or 'NOME' columns in the online spreadsheet. Ensure your sheet has these columns.")

            records = []
            
            for row in data[1:]:
                if len(row) > col_id and len(row) > col_name:
                    val_id = row[col_id]
                    val_name = row[col_name]
                    
                    if val_id and str(val_id).strip() != "":
                        clean_name = val_name if val_name else "SEM NOME"
                        records.append(Entity(id=str(val_id).strip(), name=str(clean_name).strip()))
                        
            return records
            
        except Exception as e:
            raise Exception(f"Failed to read master spreadsheet from the cloud: {e}")

    def open_spreadsheets(self, mode: int):
        self.mode = mode
        entity_name = getattr(config, 'MODE_1_NAME', "Mode_1") if mode == 1 else getattr(config, 'MODE_2_NAME', "Mode_2")
        
        if mode == 1:
            target_columns = getattr(config, 'TARGET_MODE_1_COLUMNS', [])
            allowed_sheets = self.allowed_mode_1_sheets
        else:
            target_columns = getattr(config, 'TARGET_MODE_2_COLUMNS', [])
            allowed_sheets = self.allowed_mode_2_sheets
        
        print(f"\n>>> LOCAL SCANNER ACTIVATED: Scanning allowed sheets for {entity_name}...\n")
        
        excluded_keywords = getattr(config, 'EXCLUDED_FILE_KEYWORDS', ["~", "ATUALIZADO", "LIMPO"])
        
        files_in_folder = []
        for root, dirs, files in os.walk("."):
            if ".venv" in root or "__pycache__" in root or ".git" in root or "saida" in root:
                continue
            for f in files:
                if f.endswith(".xlsx") and not any(kw in f for kw in excluded_keywords):
                    files_in_folder.append(os.path.join(root, f))
        
        for file in files_in_folder:
            try:
                wb = openpyxl.load_workbook(file, data_only=False)
                self.workbooks[file] = wb
                
                for sheet_name in wb.sheetnames:
                    if any(allowed.upper() in sheet_name.upper() for allowed in allowed_sheets):
                        sheet = wb[sheet_name]
                        
                        header = []
                        for cell in sheet[1]:
                            header.append(str(cell.value).strip().upper() if cell.value else "")
                        
                        for col_idx, col_name in enumerate(header):
                            if any(target.upper() in col_name for target in target_columns):
                                self.mapping.append({
                                    'file': file,
                                    'sheet': sheet_name,
                                    'col_idx': col_idx + 1,
                                    'col_name': col_name
                                })
            except Exception as e:
                print(f"[WARNING] Ignoring local file {file} due to read error.")

    def count_transactions(self):
        count_map = defaultdict(lambda: defaultdict(int))
        
        for mapping in self.mapping:
            wb = self.workbooks[mapping['file']]
            sheet = wb[mapping['sheet']]
            col_idx = mapping['col_idx']
            
            context_name = f"{mapping['file'].replace('.xlsx','')} | {mapping['sheet']} | {mapping.get('col_name', 'Desconhecida')}"
            
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        val_str = str(val).strip()
                        count_map[val_str][context_name] += 1
                        
        return count_map

    def apply_id_updates(self, migrations):
        migration_map = {m.source_id: m.target_id for m in migrations}
        
        for mapping in self.mapping:
            wb = self.workbooks[mapping['file']]
            sheet = wb[mapping['sheet']]
            col_idx = mapping['col_idx']
            
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        val_str = str(val).strip()
                        if val_str in migration_map:
                            cell.value = migration_map[val_str]

    def validate_migrations(self, migrations):
        problematic_ids = [m.source_id for m in migrations]
        failures = []
        
        for mapping in self.mapping:
            wb = self.workbooks[mapping['file']]
            sheet = wb[mapping['sheet']]
            col_idx = mapping['col_idx']
            
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        val_str = str(val).strip()
                        if val_str in problematic_ids:
                            failures.append(f"FAILURE: ID '{val_str}' still exists in {mapping['file']} -> {mapping['sheet']} (Row {cell.row})")
                            
        return failures

    def save_workbooks(self):
        import os
        saved_files = []
        
        output_folder = "saida"
        os.makedirs(output_folder, exist_ok=True)

        for original_path, wb in self.workbooks.items():
            file_name = os.path.basename(original_path)
            
            for kw in getattr(config, 'EXCLUDED_FILE_KEYWORDS', []):
                file_name = file_name.replace(f"{kw}_", "")
            
            new_path = os.path.join(output_folder, file_name)
                
            wb.save(new_path)
            saved_files.append(new_path)
            
        return saved_files